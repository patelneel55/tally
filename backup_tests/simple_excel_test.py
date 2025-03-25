"""
Simple Excel Export Test
-----------------------

This script tests basic Excel export functionality using openpyxl directly.
It creates a simple financial statement in Excel format.
"""

import os
from pathlib import Path
from datetime import datetime
from enum import Enum
import logging

# Set up basic logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Create output directory
output_dir = Path("analysis_results/excel_exports")
os.makedirs(output_dir, exist_ok=True)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Define simple statement type enum
    class StatementType(str, Enum):
        INCOME = "Income Statement"
        BALANCE = "Balance Sheet"
        CASH_FLOW = "Cash Flow"
    
    def create_simple_excel(output_file):
        """Create a simple Excel file with financial data."""
        logger.info(f"Creating Excel file: {output_file}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Data"
        
        # Add title
        ws["A1"] = "AAPL - Financial Data"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Add headers
        headers = ["Metric", "2023", "2022", "YoY Change"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", 
                                   end_color="D9D9D9", 
                                   fill_type="solid")
        
        # Add data
        data = [
            ["Revenue", 1000000, 900000, "=B4/C4-1"],
            ["Cost of Revenue", 600000, 550000, "=B5/C5-1"],
            ["Gross Profit", 400000, 350000, "=B6/C6-1"],
            ["Operating Expenses", 250000, 220000, "=B7/C7-1"],
            ["Operating Income", 150000, 130000, "=B8/C8-1"],
            ["Net Income", 120000, 100000, "=B9/C9-1"]
        ]
        
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                
                # Format percentages in the last column
                if col_idx == 4:
                    cell.number_format = "0.00%"
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Excel file created successfully: {output_file}")
        return True
    
    if __name__ == "__main__":
        try:
            output_file = output_dir / "simple_financial_test.xlsx"
            success = create_simple_excel(output_file)
            
            if success:
                logger.info("Test completed successfully!")
                logger.info(f"Excel file saved to: {output_file}")
            else:
                logger.error("Test failed to complete.")
        except Exception as e:
            logger.error(f"Error during test: {str(e)}", exc_info=True)

except ImportError as e:
    logger.error(f"Failed to import required modules: {str(e)}")
    logger.error("Please ensure all dependencies are installed:")
    logger.error("pip install openpyxl pandas") 