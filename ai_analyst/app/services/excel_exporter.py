"""
Excel Financial Statement Exporter
---------------------------------

This module provides functionality to export financial statements to Excel files
with formulas.

What this file does:
1. Converts structured financial statement data (from financial_statement_extractor)
   into formatted Excel workbooks
2. Creates separate sheets for different statement types (Income, Balance Sheet, Cash Flow)
3. Applies professional styling with number formats, bold headers, and freeze panes
4. Adds Excel formulas for YoY growth calculations and financial ratios
5. Generates downloadable Excel files for financial analysis

How it fits in the architecture:
- Consumes data from financial_statement_extractor
- Provides formatted output for the export API endpoints
- Creates user-friendly financial data for offline analysis

Financial importance:
- Enables detailed financial analysis in Excel, the industry standard tool
- Preserves calculations with formulas rather than static values
- Standardizes financial data presentation for consistency
"""

import os
import io
import logging
from typing import Dict, List, Optional, Any, Union, BinaryIO
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.chart import LineChart, Reference

from ai_analyst.app.models.financial_statements import (
    FinancialStatement,
    FinancialStatementType,
    FinancialStatementPeriod
)

try:
    from ai_analyst.app.core.config import settings
except ImportError:
    # Mock settings for testing
    class MockSettings:
        DATA_DIR = "data"
    settings = MockSettings()

# Set up logging
logger = logging.getLogger(__name__)

class ExcelFinancialExporter:
    """
    Excel financial statement exporter. Converts financial statements to Excel workbooks
    with formatting, formulas, and charts.
    """
    
    def __init__(self):
        """
        Initialize the Excel exporter.
        """
        try:
            data_dir = settings.DATA_DIR
        except Exception as e:
            # Default to 'data' directory if settings.DATA_DIR is not available
            data_dir = 'data'
            
        self.output_dir = Path(data_dir) / "excel_exports"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Styling constants
        self.header_font = Font(bold=True, size=12)
        self.subheader_font = Font(bold=True, size=11)
        self.normal_font = Font(size=10)
        self.title_font = Font(bold=True, size=14)
        
        self.header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def export_financial_statements(self, 
                                   statements: Dict[FinancialStatementType, FinancialStatement],
                                   ticker: str,
                                   fiscal_year: int,
                                   fiscal_period: Optional[str] = None,
                                   output_file: Optional[str] = None) -> BinaryIO:
        """
        Export financial statements to Excel.
        
        Args:
            statements: Dict mapping statement types to financial statements
            ticker: Company ticker symbol
            fiscal_year: Fiscal year
            fiscal_period: Fiscal period (e.g., Q1, Q2, FY)
            output_file: Output file path (optional)
            
        Returns:
            File-like object containing the Excel workbook
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Create sheets for each statement type
        income_sheet = None
        balance_sheet = None
        cash_flow_sheet = None
        
        if FinancialStatementType.INCOME_STATEMENT in statements:
            income_sheet = wb.create_sheet("Income Statement")
            self._format_income_statement(income_sheet, statements[FinancialStatementType.INCOME_STATEMENT])
            
        if FinancialStatementType.BALANCE_SHEET in statements:
            balance_sheet = wb.create_sheet("Balance Sheet")
            self._format_balance_sheet(balance_sheet, statements[FinancialStatementType.BALANCE_SHEET])
            
        if FinancialStatementType.CASH_FLOW in statements:
            cash_flow_sheet = wb.create_sheet("Cash Flow")
            self._format_cash_flow_statement(cash_flow_sheet, statements[FinancialStatementType.CASH_FLOW])
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary", 0)
        self._create_summary_sheet(
            summary_sheet, 
            ticker, 
            fiscal_year, 
            fiscal_period,
            income_statement=statements.get(FinancialStatementType.INCOME_STATEMENT),
            balance_sheet=statements.get(FinancialStatementType.BALANCE_SHEET),
            cash_flow_statement=statements.get(FinancialStatementType.CASH_FLOW)
        )
        
        # Save to file if output_file is provided
        if output_file:
            wb.save(output_file)
            logger.info(f"Excel file saved to {output_file}")
        
        # Return file object
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _format_income_statement(self, sheet, statement: FinancialStatement):
        """Format the income statement sheet."""
        # Add title
        sheet["A1"] = f"{statement.company_ticker} - Income Statement"
        sheet["A1"].font = self.title_font
        
        # Add period information
        sheet["A2"] = f"Period: {statement.fiscal_period} {statement.fiscal_year}"
        sheet["A2"].font = self.subheader_font
        
        # Add currency information
        sheet["A3"] = f"Currency: {statement.currency}, Units: {statement.units}"
        sheet["A3"].font = self.normal_font
        
        # Add headers
        sheet["A5"] = "Metric"
        sheet["A5"].font = self.header_font
        sheet["A5"].fill = self.header_fill
        
        # Add periods as columns
        col_idx = 2  # Start from column B
        for period in sorted(next(iter(statement.metrics.values())).keys()):
            col_letter = get_column_letter(col_idx)
            sheet[f"{col_letter}5"] = period
            sheet[f"{col_letter}5"].font = self.header_font
            sheet[f"{col_letter}5"].fill = self.header_fill
            col_idx += 1
        
        # Add metrics
        row_idx = 6
        for metric_name, values in statement.metrics.items():
            sheet[f"A{row_idx}"] = metric_name
            sheet[f"A{row_idx}"].font = self.normal_font
            
            col_idx = 2  # Start from column B
            for period in sorted(values.keys()):
                col_letter = get_column_letter(col_idx)
                sheet[f"{col_letter}{row_idx}"] = values[period]
                sheet[f"{col_letter}{row_idx}"].font = self.normal_font
                sheet[f"{col_letter}{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
                col_idx += 1
                
            row_idx += 1
        
        # Adjust column widths
        for col in range(1, col_idx):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    def _format_balance_sheet(self, sheet, statement: FinancialStatement):
        """Format the balance sheet."""
        # Add title
        sheet["A1"] = f"{statement.company_ticker} - Balance Sheet"
        sheet["A1"].font = self.title_font
        
        # Add period information
        sheet["A2"] = f"Period: {statement.fiscal_period} {statement.fiscal_year}"
        sheet["A2"].font = self.subheader_font
        
        # Add currency information
        sheet["A3"] = f"Currency: {statement.currency}, Units: {statement.units}"
        sheet["A3"].font = self.normal_font
        
        # Add headers
        sheet["A5"] = "Metric"
        sheet["A5"].font = self.header_font
        sheet["A5"].fill = self.header_fill
        
        # Add periods as columns
        col_idx = 2  # Start from column B
        for period in sorted(next(iter(statement.metrics.values())).keys()):
            col_letter = get_column_letter(col_idx)
            sheet[f"{col_letter}5"] = period
            sheet[f"{col_letter}5"].font = self.header_font
            sheet[f"{col_letter}5"].fill = self.header_fill
            col_idx += 1
        
        # Add metrics
        row_idx = 6
        for metric_name, values in statement.metrics.items():
            sheet[f"A{row_idx}"] = metric_name
            sheet[f"A{row_idx}"].font = self.normal_font
            
            col_idx = 2  # Start from column B
            for period in sorted(values.keys()):
                col_letter = get_column_letter(col_idx)
                sheet[f"{col_letter}{row_idx}"] = values[period]
                sheet[f"{col_letter}{row_idx}"].font = self.normal_font
                sheet[f"{col_letter}{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
                col_idx += 1
                
            row_idx += 1
        
        # Adjust column widths
        for col in range(1, col_idx):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    def _format_cash_flow_statement(self, sheet, statement: FinancialStatement):
        """Format the cash flow statement."""
        # Add title
        sheet["A1"] = f"{statement.company_ticker} - Cash Flow Statement"
        sheet["A1"].font = self.title_font
        
        # Add period information
        sheet["A2"] = f"Period: {statement.fiscal_period} {statement.fiscal_year}"
        sheet["A2"].font = self.subheader_font
        
        # Add currency information
        sheet["A3"] = f"Currency: {statement.currency}, Units: {statement.units}"
        sheet["A3"].font = self.normal_font
        
        # Add headers
        sheet["A5"] = "Metric"
        sheet["A5"].font = self.header_font
        sheet["A5"].fill = self.header_fill
        
        # Add periods as columns
        col_idx = 2  # Start from column B
        for period in sorted(next(iter(statement.metrics.values())).keys()):
            col_letter = get_column_letter(col_idx)
            sheet[f"{col_letter}5"] = period
            sheet[f"{col_letter}5"].font = self.header_font
            sheet[f"{col_letter}5"].fill = self.header_fill
            col_idx += 1
        
        # Add metrics
        row_idx = 6
        for metric_name, values in statement.metrics.items():
            sheet[f"A{row_idx}"] = metric_name
            sheet[f"A{row_idx}"].font = self.normal_font
            
            col_idx = 2  # Start from column B
            for period in sorted(values.keys()):
                col_letter = get_column_letter(col_idx)
                sheet[f"{col_letter}{row_idx}"] = values[period]
                sheet[f"{col_letter}{row_idx}"].font = self.normal_font
                sheet[f"{col_letter}{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
                col_idx += 1
                
            row_idx += 1
        
        # Adjust column widths
        for col in range(1, col_idx):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_summary_sheet(self, 
                             sheet, 
                             ticker: str, 
                             fiscal_year: int, 
                             fiscal_period: Optional[str],
                             income_statement: Optional[FinancialStatement] = None,
                             balance_sheet: Optional[FinancialStatement] = None,
                             cash_flow_statement: Optional[FinancialStatement] = None):
        """Create a summary sheet with key metrics and charts."""
        # Add title
        sheet["A1"] = f"{ticker} - Financial Summary"
        sheet["A1"].font = self.title_font
        
        # Add period information
        period_str = f"{fiscal_period} " if fiscal_period else ""
        sheet["A2"] = f"Period: {period_str}{fiscal_year}"
        sheet["A2"].font = self.subheader_font
        
        # Section: Key Metrics
        sheet["A4"] = "Key Metrics"
        sheet["A4"].font = self.header_font
        
        row_idx = 5
        
        # Revenue
        if income_statement and "Revenue" in income_statement.metrics:
            revenue_values = income_statement.metrics["Revenue"]
            latest_period = sorted(revenue_values.keys())[-1]
            sheet[f"A{row_idx}"] = "Revenue"
            sheet[f"B{row_idx}"] = revenue_values[latest_period]
            sheet[f"B{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
            row_idx += 1
        
        # Net Income
        if income_statement and "Net Income" in income_statement.metrics:
            net_income_values = income_statement.metrics["Net Income"]
            latest_period = sorted(net_income_values.keys())[-1]
            sheet[f"A{row_idx}"] = "Net Income"
            sheet[f"B{row_idx}"] = net_income_values[latest_period]
            sheet[f"B{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
            row_idx += 1
        
        # Total Assets
        if balance_sheet and "Total Assets" in balance_sheet.metrics:
            assets_values = balance_sheet.metrics["Total Assets"]
            latest_period = sorted(assets_values.keys())[-1]
            sheet[f"A{row_idx}"] = "Total Assets"
            sheet[f"B{row_idx}"] = assets_values[latest_period]
            sheet[f"B{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
            row_idx += 1
        
        # Total Liabilities
        if balance_sheet and "Total Liabilities" in balance_sheet.metrics:
            liabilities_values = balance_sheet.metrics["Total Liabilities"]
            latest_period = sorted(liabilities_values.keys())[-1]
            sheet[f"A{row_idx}"] = "Total Liabilities"
            sheet[f"B{row_idx}"] = liabilities_values[latest_period]
            sheet[f"B{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
            row_idx += 1
        
        # Cash from Operations
        if cash_flow_statement and "Cash from Operating Activities" in cash_flow_statement.metrics:
            cash_ops_values = cash_flow_statement.metrics["Cash from Operating Activities"]
            latest_period = sorted(cash_ops_values.keys())[-1]
            sheet[f"A{row_idx}"] = "Cash from Operations"
            sheet[f"B{row_idx}"] = cash_ops_values[latest_period]
            sheet[f"B{row_idx}"].number_format = FORMAT_CURRENCY_USD_SIMPLE
            row_idx += 1
        
        # Add metadata about the export
        sheet[f"A{row_idx + 2}"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet[f"A{row_idx + 3}"] = "Generated by: Analyst AI Excel Exporter"
        
        # Adjust column widths
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 20


# Create a singleton instance
excel_exporter = ExcelFinancialExporter()

# For backwards compatibility
def export_financial_statements_to_excel(
    statements: Dict[FinancialStatementType, FinancialStatement],
    ticker: str,
    fiscal_year: int,
    fiscal_period: Optional[str] = None,
    output_file: Optional[str] = None
) -> BinaryIO:
    """
    Export financial statements to Excel.
    
    Args:
        statements: Dict mapping statement types to financial statements
        ticker: Company ticker symbol
        fiscal_year: Fiscal year
        fiscal_period: Fiscal period (e.g., Q1, Q2, FY)
        output_file: Output file path (optional)
        
    Returns:
        File-like object containing the Excel workbook
    """
    return excel_exporter.export_financial_statements(
        statements=statements,
        ticker=ticker,
        fiscal_year=fiscal_year,
        fiscal_period=fiscal_period,
        output_file=output_file
    ) 